# # model/excel_manager.py

# import csv
# import os

# def save_to_csv(data_list, save_path):
#     """
#     테이블 데이터를 CSV로 저장
#     썸네일은 경로만 텍스트로 저장됨
#     """
#     headers = ["Thumbnail", "Roll", "Shot Name", "Version", "Type", "Path"]

#     with open(save_path, mode='w', newline='', encoding='utf-8') as f:
#         writer = csv.DictWriter(f, fieldnames=headers)
#         writer.writeheader()
#         for item in data_list:
#             writer.writerow({
#                 "Thumbnail": item["thumbnail"], 
#                 "Roll": item["roll"],
#                 "Shot Name": item["shot_name"],
#                 "Version": item["version"],
#                 "Type": item["type"],
#                 "Path": item["path"],
#             })

#     print(f" CSV 저장 완료: {save_path}")


import os
import re
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as XLImage

def get_next_versioned_filename(base_path, prefix="scanlist", ext=".xlsx"):
    """
    기존 파일을 기준으로 다음 버전명을 자동 생성
    예: scanlist_v001.xlsx → scanlist_v002.xlsx
    """
    dir_name = os.path.dirname(base_path)
    base_name = os.path.splitext(os.path.basename(base_path))[0]

    # 파일명에서 prefix_v### 형식 추출
    pattern = re.compile(rf"{re.escape(prefix)}_v(\d{{3}})")
    existing_versions = []

    for f in os.listdir(dir_name):
        match = pattern.match(f)
        if match:
            existing_versions.append(int(match.group(1)))

    next_version = max(existing_versions, default=0) + 1
    return os.path.join(dir_name, f"{prefix}_v{next_version:03d}{ext}")

def save_to_excel(table_widget, save_base_path):
    """
    PySide 테이블 위젯 데이터를 .xlsx 파일로 자동 버전명 붙여 저장
    """
    from openpyxl import Workbook

    save_path = get_next_versioned_filename(save_base_path)

    wb = Workbook()
    ws = wb.active
    ws.title = "ScanList" 

    # 헤더 저장
    headers = ["Thumbnail", "Roll", "Shot Name", "Version", "Type", "Path"]
    for col in range(table_widget.columnCount()):
        header = table_widget.horizontalHeaderItem(col).text()
        headers.append(header)
    ws.append(headers)


    for row_data in range(table_widget.rowCount()):
        img_path = row_data["thumbnail"]
        row = [
            "",  # 썸네일은 나중에 이미지로 채움
            row_data["roll"],
            row_data["shot_name"],
            row_data["version"],
            row_data["type"],
            row_data["path"],
        ]
        ws.append(row)

        if img_path and os.path.exists(img_path):
            img = XLImage(img_path)
            img.width = 100
            img.height = 60
            cell = f"A{ws.max_row}"  # 현재 행의 A열에 이미지 삽입
            ws.add_image(img, cell)
    # 내용 저장
    for row in range(table_widget.rowCount()):
        row_data = []
        for col in range(table_widget.columnCount()):
            item = table_widget.item(row, col)
            if item:
                row_data.append(item.text())
            else:
                row_data.append("")  # 빈 칸 처리
        ws.append(row_data)

    wb.save(save_path)
    print(f" 엑셀 저장 완료: {save_path}")



def load_excel_data(xlsx_path):
    """
    저장된 .xlsx 파일을 열어 딕셔너리 리스트로 반환
    """
    wb = load_workbook(xlsx_path)
    ws = wb.active

    headers = [cell.value for cell in ws[1]]
    data_list = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        row_data = dict(zip(headers, row))
        data_list.append(row_data)

    return data_list